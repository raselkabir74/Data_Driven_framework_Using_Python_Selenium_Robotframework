"""Base Page implementations"""
import openpyxl
import xlrd
from openpyxl.utils import get_column_letter

import TestFramework.Libraries.Modules.browser as Browser
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, ElementNotVisibleException, WebDriverException
import datetime
import time
import random
import string
import win32com.client as client
import os.path
import glob
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select


class BasePage:
    # Start: Base page locators
    default_window_identifier_locator = (By.CLASS_NAME, "logout")
    calendar_header_locator = (By.XPATH, "//kendo-calendar-header[@class='k-calendar-header']/span")
    calendar_next_arrow_button_locator = (By.XPATH, "//button[@title='Navigate to next view']/span")
    calendar_previous_arrow_button_locator = (By.XPATH, "//button[@title='Navigate to previous view']/span")
    today_button_locator = (By.XPATH, "//span[normalize-space(text())='TODAY']")
    biddle_logo_locator = (By.XPATH, "//div[@class='logo']/img")
    # End: Base page locators

    def wait(self, value=None):
        """
        Implementing wait functionality
        :param value: Seconds
        :return:WebDriver Wait instance
        """
        return Browser.wait(value)

    def hover(self, element):
        """
        Implementing hover functionality
        :param element: WebElement locator
        :return:Mouse hover on WebElement
        """
        Browser.hover(element)

    def maximize_window(self):
        """
        Implementing maximize window functionality
        Maximizes the current window that webdriver is using
        :return: Window Maximize
        """
        Browser.maximize_window()

    def refresh_window(self):
        """
        Implementing refresh window functionality
        :return: Refreshes the current page.
        """
        Browser.refresh_window()

    def close_browser(self):
        """
        Implementing close browser functionality
        :return: Closes the current window.
        """
        Browser.close_browser()

    def page_title(self):
        """
        Implementing page title functionality
        Returns the title of the current page.
        :return: Current Page Title
        """
        return Browser.page_title()

    def script_executor_click(self, element, execute_async_script=False):
        """
        Implementing script executor click functionality
        :param element: WebElement locator
        :param execute_async_script:
        :return: Click on WebElement using JavaScript Executor,
        """
        Browser.script_executor_click(element, execute_async_script)

    def script_executor(self, element, script=None):
        """
        Implementing script executor functionality
        :param element: WebElement
        :return: Click on WebElement using JavaScript Executor
        """
        return Browser.script_executor(element, script)

    def find_element(self, locator, time_out=Browser.get_time_out_value(), to_be_present=True, to_be_clickable=False, to_be_visible=False):
        """
        Implementing find element functionality
        :param locator:
        :param time_out:
        :param to_be_present:
        :param to_be_clickable:
        :param to_be_visible:
        :return:
        """
        element = None
        if to_be_present is True:
            element = self.wait(time_out).until(EC.presence_of_element_located(locator), 'locator not found before specified time out')
        elif to_be_clickable is True:
            element = self.wait(time_out).until(EC.element_to_be_clickable(locator), 'locator not clickable before specified time out')
        elif to_be_visible is True:
            element = self.wait(time_out).until(EC.visibility_of_element_located(locator), 'locator not visible before specified time out')
        return element

    def is_element_present(self, locator, time_out=Browser.get_time_out_value()):
        """
        Implementing is element present functionality
        Verify WebElement locator in page/ui
        :param locator: WebElement locator
        :return: True/False
        """
        is_present = None
        try:
            self.wait(time_out).until(EC.presence_of_element_located(locator), 'locator not found before specified time out')
            is_present = True
        except:
            is_present = False
        finally:
            return is_present

    def switch_to_default_content(self):
        """
        Implementing switch to default content functionality
        :return:
        """
        Browser.switch_to_default_content()

    def wait_for_ajax_spinner_load(self, timeout=60, ajax_spinner_locator=None, myaap_screen=False):
        """
        Implementing wait for ajax spinner load functionality
        Wait until Ajax Spinner exists on UI. Once it is disappear, exit from loop
        Default time out value up to 60 secs
        :param timeout:
        :param ajax_spinner_locator:
        :param myaap_screen:
        :return:
        """
        try:
            if ajax_spinner_locator is not None:
                ajax_spinner_load_locator = ajax_spinner_locator
            else:
                ajax_spinner_load_locator = (By.XPATH, "//div[@class='lds-ellipsis']")
            ajax_spinner_control = self.wait(2).until(EC.visibility_of_all_elements_located(ajax_spinner_load_locator), 'ajax spinner load locator not found before specified time out')
            end_time = time.time() + timeout
            flag = False
            while len(ajax_spinner_control) > 0:
                if myaap_screen is True:
                    ajax_spinner_control = self.wait(2).until(EC.visibility_of_all_elements_located(ajax_spinner_load_locator), 'ajax spinner load locator not found before specified time out')
                    for element in ajax_spinner_control:
                        if "display: none;" in element.get_attribute("style"):
                            flag = True
                            break
                else:
                    ajax_spinner_control = self.wait(2).until(EC.visibility_of_all_elements_located(ajax_spinner_load_locator), 'ajax spinner load locator not found before specified time out')
                if flag:
                    break
                if time.time() > end_time:
                    break

        except TimeoutException:
            pass
        except NoSuchElementException:
            pass
        except ElementNotVisibleException:
            pass
        except WebDriverException:
            pass

    def get_current_date(self):
        """
        Implementing get current date functionality
        :return: Current date
        """
        return time.strftime("%m/%d/%Y")

    def get_future_date(self):
        """
        Implementing get future date functionality
        :return: future date
        """
        current_date = self.get_current_date()
        modified_date = datetime.datetime.strptime(current_date, "%m/%d/%Y")
        new_date = str(modified_date + datetime.timedelta(days=7))
        return new_date

    def random_string_generator(self, size=6, chars=string.ascii_uppercase + string.digits):
        """
        Implementing random string generator functionality
        :param size:
        :param chars:
        :return: random string
        """
        return ''.join(random.SystemRandom().choice(chars) for _ in range(size))

    def set_value_into_input_field(self, input_field_locator, value, use_win32com=False, script_executor=False):
        """
        Implementing set value into input field functionality
        :param input_field_locator:
        :param value:
        :param use_win32com:
        :param script_executor:
        :return:
        """
        input_field_element = self.wait().until(EC.presence_of_element_located(input_field_locator), 'input field locator not found before specified time out')
        if use_win32com is True:
            if script_executor is True:
                self.script_executor_click(input_field_element)
            else:
                self.click_element(input_field_locator, wait_after_click=False)
            input_field = client.Dispatch("WScript.Shell")
            input_field.SendKeys("^a")
            input_field.SendKeys("{DEL}")
            input_field.SendKeys(str(value))
        else:
            input_field_element.clear()
            input_field_element.send_keys(value)
        self.wait_for_ajax_spinner_load()

    def click_element(self, locator, script_executor=False, hover=False, error_message=None, wait_after_click=True, scroll_into_view=True):
        """
        Implementing click element functionality
        :param locator:
        :param script_executor:
        :param hover:
        :param error_message:
        :param wait_after_click:
        :param scroll_into_view:
        :return:
        """
        if error_message is None or error_message == "":
            error_message = 'locator not found before specified time out'
        element = self.wait().until(EC.presence_of_element_located(locator), error_message)
        if hover is True:
            self.hover(element)
        if scroll_into_view:
            self.scroll_into_view(element)
        if script_executor is True:
            self.script_executor_click(element)
        else:
            # TODO: script_executor_click() has been commented out intentionally. After observing the result the statement will be reverted back.
            element.click()
            # self.script_executor_click(element)
        if wait_after_click:
            self.wait_for_ajax_spinner_load()

    def drag_and_drop(self, source, target, with_action_chain=False, with_custom_action_chain=False):
        """
        Implementing drag and drop functionality
        :param source:
        :param target:
        :param with_action_chain:
        :param with_custom_action_chain:
        :return:
        """
        Browser.drag_and_drop(source, target, with_action_chain, with_custom_action_chain)

    def get_date(self, given_date=None, current_date=False, future_date=False, past_date=False, number_of_days_to_add=7, number_of_days_to_back=7, first_day_of_last_month=False, last_day_of_last_month=False, first_day_of_current_month=False, last_day_of_current_month=False):
        """
        Implementing get date functionality.
            Current day is set to given_date if any date parameter is passed, Returns current date if current_date is True, future date if future_date is True, number_of_days_to_add indicates the number of day(s) to add with current date to get the future date,
            Returns first day of previous month if first_day_of_last_month is True, last day of previous month if last_day_of_last_month is True, first day of current month if first_day_of_current_month is True, last day of current month if last_day_of_current_month is True.
        :param given_date:
        :param current_date:
        :param future_date:
        :param past_date:
        :param number_of_days_to_add:
        :param number_of_days_to_back:
        :param first_day_of_last_month:
        :param last_day_of_last_month:
        :param first_day_of_current_month:
        :param last_day_of_current_month:
        :return: expected_date
        """
        if given_date is None:
            current_day = datetime.date.today()
        else:
            current_day = datetime.datetime.strptime(given_date, "%m/%d/%Y")
        expected_date = current_day
        if current_date is True:
            expected_date = "%d/%d/%d" % (current_day.month, current_day.day, current_day.year)
        elif future_date is True:
            future_day = current_day + datetime.timedelta(days=number_of_days_to_add)
            expected_date = "%d/%d/%d" % (future_day.month, future_day.day, future_day.year)
        elif past_date is True:
            past_day = current_day - datetime.timedelta(days=number_of_days_to_back)
            expected_date = "%d/%d/%d" % (past_day.month, past_day.day, past_day.year)
        elif last_day_of_current_month is True:
            next_month = current_day.replace(day=28) + datetime.timedelta(days=4)
            last_day_of_current_month = next_month - datetime.timedelta(days=next_month.day)
            expected_date = "%d/%d/%d" % (
            last_day_of_current_month.month, last_day_of_current_month.day, last_day_of_current_month.year)
        elif first_day_of_current_month is True:
            first_day_of_current_month = current_day.replace(day=1)
            expected_date = "%d/%d/%d" % (
            first_day_of_current_month.month, first_day_of_current_month.day, first_day_of_current_month.year)
        elif last_day_of_last_month is True:
            last_day_of_last_month = current_day - datetime.timedelta(days=current_day.day)
            expected_date = "%d/%d/%d" % (
            last_day_of_last_month.month, last_day_of_last_month.day, last_day_of_last_month.year)
        elif first_day_of_last_month is True:
            last_month = current_day - datetime.timedelta(days=current_day.day)
            last_month_first_day = last_month.replace(day=1)
            expected_date = "%d/%d/%d" % (
            last_month_first_day.month, last_month_first_day.day, last_month_first_day.year)
        return expected_date

    def switch_to_frame(self, frame_locator):
        """
        Implementing switch to frame functionality
        :param frame_locator:
        :return:
        """
        frame_element = self.wait().until(EC.presence_of_element_located(frame_locator), 'frame locator not found before specified timeout')
        Browser.switch_to_frame(frame_element)

    def get_text_from_element(self, locator, is_a_input_field=False):
        """
        Implementing get text from element functionality
        :return: element's text
        :param locator:
        :param is_a_input_field:
        """
        element = self.wait().until(EC.presence_of_element_located(locator), 'locator not found before specified time out')
        if is_a_input_field is True:
            text = str(element.get_attribute("value")).strip()
        else:
            text = str(element.text).strip()
        return text

    def get_attribute_value_of_element_locator(self, element_locator, attribute_name):
        """
        Implementing get attribute value from element locator functionality
        :param element_locator: element_locator
        :param attribute_name: attribute_name
        :return: attribute_value
        """
        element = self.wait().until(EC.presence_of_element_located(element_locator), 'locator not found before specified time out')
        attribute_value = element.get_attribute(attribute_name)
        return attribute_value

    def goto_specific_url(self, value):
        """
        Implementing go to specific url functionality
        :param value:
        :return:
        """
        Browser.goto_specific_url(value)

    def get_current_browser_name(self):
        """
        Implementing get current browser name functionality
        :return: _current_browser_name
        """
        return Browser.get_current_browser_name()

    def is_element_visible(self, locator, timeout=30):
        """
        Implementing is element visible functionality
        Verify WebElement locator is visible in page/ui
        :param locator: WebElement locator
        :return: True/False
        """
        # To Do: Error log is not writing in the report. We have to work on this
        is_visible = None
        try:
            self.wait(timeout).until(EC.visibility_of_element_located(locator), 'locator not found before specified time out')
            is_visible = True
        except:
            is_visible = False
        return is_visible

    def is_element_clickable(self, locator, timeout=30):
        """
        Implementing is element clickable functionality
        Verify WebElement locator is clickable in page/ui
        :param locator: WebElement locator
        :return: True/False
        """
        # To Do: Error log is not writing in the report. We have to work on this
        is_clickable = None
        try:
            self.wait(timeout).until(EC.element_to_be_clickable(locator), 'locator not found before specified time out')
            is_clickable = True
        except:
            is_clickable = False
        return is_clickable

    def is_expected_file_present_in_directory(self, file_path):
        """
        Implementing is expected file present in directory functionality
        :return: True/False
        """
        modified_file_path = file_path.replace('/', '\\')
        return os.path.isfile(modified_file_path)

    def double_click(self, locator):
        """
        Implementing double click functionality
        :param locator:
        :return:
        """
        custom_error_message = 'locator not found before specified time out'
        element = self.wait().until(EC.element_to_be_clickable(locator), custom_error_message)
        Browser.double_click(element)

    def get_current_page_url(self):
        """
        Implementing get current page url functionality
        :return: current_url
        """
        return Browser.get_current_page_url()

    def get_time_out_value(self):
        """
        Implementing get time out value functionality
        :return: time out value
        """
        return Browser.get_time_out_value()

    def set_existing_handles(self):
        """
        Implementing set existing handles functionality
        :return:
        """
        Browser.set_existing_handles()

    def clear_existing_handles(self):
        """
        Implementing clear existing handles functionality
        :return:
        """
        Browser.clear_existing_handles()

    def switch_to_previous_window(self):
        """
        Implementing switch to previous window functionality
        :return:
        """
        Browser.switch_to_previous_window()

    def switch_to_default_window(self):
        """
        Implementing switch to default window functionality
        :return: Switch to default screen/window
        """
        Browser.switch_to_default_window()
        try:
            application_build_and_release_number_element = self.wait(30).until(EC.presence_of_element_located(self.default_window_identifier_locator), 'default window identifier locator not found before specified time out')
            application_build_and_release_number_element.click()
        except:
            pass

    def get_current_window_handle(self):
        """
        Implementing get current window handle functionality
        :return: _current_window_handle
        """
        return Browser.get_current_window_handle()

    def switch_to_window(self, current_handler=None):
        """
        Implementing switch to window functionality
        :param current_handler: current_handler
        :return:Switch to child screen/window
        """
        if current_handler is None:
            Browser.switch_to_window()
        else:
            Browser.switch_to_window(handler=current_handler)

    def select_date_in_calendar(self, date_textbox_locator, date, toggle_calendar=False):
        """
        Implementing select date in calender functionality
        :param date_textbox_locator:
        :param date:
        :param toggle_calendar:
        :return:
        """
        month, day, year = date.split("/")
        self.click_element(date_textbox_locator, error_message='date textbox locator not found before specified time out')
        months_list = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        if str(date) == str(self.get_date(current_date=True)):
            self.click_element(date_textbox_locator, error_message='date textbox locator not found before specified time out', scroll_into_view=False)
            self.click_element(self.today_button_locator, error_message='today button locator not found before specified time out', scroll_into_view=False)
        else:
            if toggle_calendar is True:
                calender_header_locator = (By.XPATH, "//span[@class='k-button k-bare k-title']")
                self.click_element(calender_header_locator, error_message='calendar header locator not found before specified time out', scroll_into_view=False)
                calendar_year_locator = (By.XPATH, "//kendo-calendar/kendo-calendar-navigation[@class='k-calendar-navigation']/kendo-virtualization[@class='k-content k-scrollable']/ul[@class='k-reset']/descendant::span[text()='%s']" % str(year))
                current_year = self.get_text_from_element(self.calendar_header_locator).split('-')[0]
                if current_year < year:
                    key = Keys.ARROW_DOWN
                else:
                    key = Keys.ARROW_UP

                while not self.is_element_present(calendar_year_locator, time_out=5):
                    self.press_keyboard_keys(key)

                self.click_element(calendar_year_locator, script_executor=True, error_message='calendar year locator not found before specified time out', scroll_into_view=False)
                calendar_month_locator = (By.XPATH, "//kendo-calendar-viewlist[@class='k-calendar-view k-calendar-yearview']/kendo-virtualization[@class='k-content k-scrollable']/table/tbody[2]/descendant::span[@class='k-link' and text()='%s']" % str(months_list[int(month)]))
                self.click_element(calendar_month_locator, error_message='calendar month locator not found before specified time out', scroll_into_view=False)
                calendar_day_locator = (By.XPATH, "//kendo-calendar-viewlist[@class='k-calendar-view k-calendar-monthview']/kendo-virtualization[@class='k-content k-scrollable']/table/tbody[1]/descendant::span[@class='k-link' and text()='%s']" % str(day))
                self.click_element(calendar_day_locator, error_message='calendar day locator not found before specified time out', scroll_into_view=False)
            else:
                current_year_month, _ = self.get_text_from_element(self.calendar_header_locator).split('-')
                if year not in current_year_month:
                    self.click_element(self.calendar_header_locator, error_message='calendar header locator not found before specified time out', scroll_into_view=False)
                    self.click_element(self.calendar_header_locator, error_message='calendar header locator not found before specified time out', scroll_into_view=False)
                    self.select_correct_year(year)
                    calendar_month_locator = (By.XPATH, "//td/span[text()='%s']" % str(months_list[int(month)]))
                    self.click_element(calendar_month_locator, error_message='calendar month locator not found before specified time out', scroll_into_view=False)
                else:
                    if str(months_list[int(month)]) not in current_year_month:
                        self.click_element(self.calendar_header_locator, error_message='ccalendar header locator not found before specified time out', scroll_into_view=False)
                        calendar_month_locator = (By.XPATH, "//td/span[text()='%s']" % str(format(months_list[int(month)])))
                        self.click_element(calendar_month_locator, error_message='calendar month locator not found before specified time out', scroll_into_view=False)
                calendar_day_locator = (By.XPATH, "//td/span[text()='%s']" % str(day))
                self.click_element(calendar_day_locator, error_message='calendar day locator not found before specified time out', scroll_into_view=False)
        self.click_element(self.biddle_logo_locator, error_message='biddle logo locator not found before specified time out', scroll_into_view=False)

    def select_correct_year(self, year):
        """
        Implementing select correct year functionality
        :param year:
        :return:
        """
        from_year, _, _, to_year = self.get_text_from_element(self.calendar_header_locator).split('-')
        while int(year) < int(from_year) or int(year) > int(to_year):
            if int(year) < int(from_year):
                self.click_element(self.calendar_previous_arrow_button_locator, error_message='calendar previous arrow button locator not found before specified time out', scroll_into_view=False)
            else:
                self.click_element(self.calendar_next_arrow_button_locator, error_message='calendar next arrow button locator not found before specified time out', scroll_into_view=False)
            from_year, _, _, to_year = self.get_text_from_element(self.calendar_header_locator).split('-')
        calendar_year_locator = (By.XPATH, "//span[@class='k-link' and normalize-space(text())='%s']" % str(year))
        self.click_element(calendar_year_locator, script_executor=True, error_message='calendar year locator not found before specified time out', scroll_into_view=False)

    def select_item_from_single_selection_dropdown(self, dropdown_name, item_name):
        """
        Implementing select item from single selection dropdown functionality
        :param dropdown_name:
        :param item_name:
        :return:
        """
        dropdown_arrow_locator = (By.XPATH, "//span[normalize-space(text())='%s']/../descendant::span[contains(@class, 'k-select')]" % dropdown_name)
        self.click_element(dropdown_arrow_locator, error_message='dropdown arrow locator not found before specified time out')
        item_locator = (By.XPATH, "//li[contains(text(), '%s')]" % str(item_name))
        self.click_element(item_locator, error_message='dropdown item locator not found before specified time out')

    def press_keyboard_keys(self, keys):
        """
        Implementing press keyboard keys functionality
        :param keys:
        :return:
        """
        for key in keys:
            Browser.press_keyboard_key(key)

    def select_date_manually(self, locator, date):
        """
        Implementing select security log date range manually functionality
        :param date:
        :return:
        """
        month, day, year = date.split("/")
        expected_date = month+day+year
        self.click_element(locator)
        self.click_element(locator)
        self.press_keyboard_keys(Keys.ARROW_LEFT)
        self.press_keyboard_keys(Keys.ARROW_LEFT)
        self.press_keyboard_keys(expected_date)
        self.click_element(self.biddle_logo_locator, error_message='biddle logo locator not found before specified time out')

    def clear_text_field(self, input_field_locator):
        """
        Implementing functionality to clear text filed
        :param input_field_locator: Locator of the input element to clear
        :return:
        """
        input_field_element = self.wait().until(EC.presence_of_element_located(input_field_locator), 'input field locator not found before specified time out')
        input_field_element.clear()

    def select_specific_tab(self, tab_name):
        """
        Implementing select specific tab functionality
        :param tab_name:
        :return:
        """
        tab_locator = (By.XPATH, "//a[text()='%s']" % tab_name)
        self.click_element(tab_locator, script_executor=True, error_message='tab locator not found before specified time out')

    def accept_alert(self):
        """
        Implementing accept alert functionality
        :return:
        """
        Browser.accept_alert()

    def clear_search_bar(self, search_bar_locator):
        """
        Implementing functionality to clear search field by repeated pressing of backspace
        :param search_bar_locator: locator for search bar
        :return:
        """
        input_field_element = self.wait().until(EC.presence_of_element_located(search_bar_locator), 'input field locator not found before specified time out')
        text_length = len(self.get_text_from_element(search_bar_locator, True))
        for key in range(text_length):
            input_field_element.send_keys(Keys.BACKSPACE)

    def move_to_element(self, locator):
        """
        Implementing move to element functionality
        :param locator:
        :return:
        """
        element = self.find_element(locator)
        Browser.move_to_element(element)

    def move_to_element_without_locator(self, element):
        """
        Implementing move to element functionality
        :param element:
        :return:
        """
        Browser.move_to_element(element)

    def is_download_successful(self, timestamp, file_extension=None, time_out=180):
        """
        Implementing is download successful functionality
        :param timestamp:
        :param file_extension:
        :param time_out:
        :return: True/False
        """
        max_time = time.time()+time_out
        status = False
        current_directory = os.getcwd()

        if file_extension is None:
            file_url = os.path.join(os.getcwd(), Browser._download_folder_name, '*.*')
        else:
            file_name = '*.' + file_extension.strip('.')
            file_url = os.path.join(os.getcwd(), Browser._download_folder_name, file_name)
        try:
            os.chdir(os.path.join(os.getcwd(), Browser._download_folder_name))
            while time.time() < max_time:
                file_list = glob.glob(file_url)
                latest_file = max(file_list, key=os.path.getctime)
                if os.path.getctime(latest_file) > timestamp:
                    status = True
                    break
        except:
            status = False
            raise
        finally:
            os.chdir(current_directory)
            return status

    def scroll_into_view(self, element):
        """
        Implementing scroll into view functionality
        :param element: WebElement
        """
        Browser.scroll_into_view(element)

    def is_alert_present(self, time_out=5):
        """
        Implementing is alert present functionality
        :param time_out:
        :return: True/False
        """
        status = False
        try:
            if self.wait(time_out).until(EC.alert_is_present()):
                status = True
        except:
            pass
        finally:
            return status

    def select_multiple_element(self, element):
        """
        Implementing select multiple element functionality
        :param element:
        :return:
        """
        Browser.select_multiple_element(element)

    def get_location(self, locator):
        """
        Implementing get location functionality
        :param locator:
        :return:
        """
        locator_element = self.wait(5).until(EC.presence_of_element_located(locator))
        return locator_element.location

    def get_screen_coordinates_from_browser_coordinates(self, browser_x, browser_y):
        """
        Implementing get screen coordinates from browser coordinates functionality
        :param browser_x:
        :param browser_y:
        :return:
        """
        browser_navigation_panel_height = self.script_executor('return window.outerHeight - window.innerHeight;')
        return {'x': browser_x, 'y': browser_y + browser_navigation_panel_height}

    def press_key_combination(self, locator, key1, key2):
        """
        Implementing press key combination functionality
        :param locator:
        :param key1:
        :param key2:
        :return:
        """
        locator_element = self.wait(5).until(EC.presence_of_element_located(locator))
        Browser.press_key_combination(locator_element, key1, key2)

    def select_dropdown_item(self, id, item):
        dropdown = Select(self.wait().until(EC.presence_of_element_located(id)))
        dropdown.select_by_visible_text(item)

    def read_data_from_excel_file(self, file_path, sheet_name, keyword_name):
        """
        Implementing read data from excel file functionality
        :param file_path:
        :param sheet_name:
        :return: _excel_data_dictionary
        """
        value = None
        try:

            workbook = xlrd.open_workbook(file_path)
            worksheet = workbook.sheet_by_name(sheet_name)
            for row in range(worksheet.nrows):
                for column in range(worksheet.ncols):
                    if worksheet.cell(row, column).value == keyword_name:
                        value = worksheet.cell(row, column+1).value
                        break
        except IOError:
            raise
        finally:
            return str(value)

    def write_test_data(self, file_path, sheet_name, keyword_name, new_value):
        """
        Implementing returning test data for a test suite
        :param file_path:
        :param sheet_name:
        :param column:
        :param value:
        :return:
        """
        try:
            row_index = 0
            column_index = 0
            workbook = xlrd.open_workbook(file_path)
            worksheet = workbook.sheet_by_name(sheet_name)
            for row in range(worksheet.nrows):
                for column in range(worksheet.ncols):
                    if worksheet.cell(row, column).value == keyword_name:
                        row_index = row + 1
                        column_index = column + 2
                        break
            column_header = get_column_letter(column_index)
            expected_cell = column_header + str(row_index)
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook[sheet_name]
            worksheet[expected_cell] = new_value
            workbook.save(file_path)
            workbook.close()
        except IOError:
            raise